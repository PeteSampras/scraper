# only some of these will be required. will clean up after.
# pip install bs4
# pip install openpyxl
# pip install requests
# pip install urllib
# pip install psutil
import requests as req
import urllib.request
import os
from bs4 import BeautifulSoup
import openpyxl
import smtplib
import configparser
import psutil

#from datetime import datetime

# pragma region globals
# set global variables for show_directory and excel file name
web_page= 'https://eztv.ag/shows/'
show_directory="f:/misc/"
wkbook_name = "Shows.xlsx" # columns: Name|EZTV ID|Season|Episode|Min Resolution|Max Resolution|Comments|Timestamp
wksheet_name = 'Shows'
only_get_last = 'last'
show_xls = show_directory+wkbook_name

# config file settings
config_file="F:/misc/config.ini"
config_header="EMAIL"                           # config file email section [header]. [EMAIL]
config_sender_email='sender_address'            # sender email address key
config_sender_password='sender_password'        # sender email password key
config_to_email='to_address'                    # recipient email address key

# email info
smtp_server='smtp.gmail.com'                    # smtp server
smtp_port=587                                   # smtp port

# config parse
config = configparser.ConfigParser()            # setup
config.read(config_file)                        # read file
if config_header in config:                     # check for all those settings
    if config_sender_email in config[config_header]:
        sender_email_address=config[config_header][config_sender_email]
    if config_sender_password in config[config_header]:
        sender_email_password=config[config_header][config_sender_password]
    if config_to_email in config[config_header]:
        to_email_address=config[config_header][config_to_email]

# email function
def email_this(subject,msg):
    s = smtplib.SMTP(smtp_server, smtp_port)                    # check server
    s.ehlo()                                                    # handshake server
    s.starttls()                                                # turn on TLS
    s.login(sender_email_address, sender_email_password)        # log in
    message = "\r\n".join([                                     # create email
            "From: " + sender_email_address,
            "To: " + to_email_address,
            "Subject: " + subject,
            msg
            ])
    s.sendmail(sender_email_address, to_email_address, message) #send email
    s.quit()                                                    # quit

# helper function to see if a number is an integer to check season data
def is_valid_int(s):
    try:
        int(s)
    except ValueError:
        return False
    else:
        return True

# helper function to get episode info
class parse_episode:
    def __init__(self,season,episode):
        self.season=season
        self.episode=episode
        self.resolution=480

def get_ep_info(my_file): # this parse is needed to get season, episode, and resolution info from the name of the file
    temp = parse_episode(0,0)
    if " " in my_file:
        split=my_file.replace('.',' ').split(' ')
    else:
        split=my_file.split('.')
    for word in split:
        if len(word)==6 and 's' in word[0].lower() and 'e' in word[3].lower():
            if is_valid_int(word[1:3]) and is_valid_int(word[4:]):
               temp = parse_episode(int(word[1:3]),int(word[4:]))
        if len(word)==10 and 's' in word[0].lower() and 'e' in word[7].lower():
            if is_valid_int(word[1:3]) and is_valid_int(word[8:]):
               temp = parse_episode(int(word[1:3]),int(word[8:]))
        if "hdtv" in word.lower() or "web" in word.lower() or "480" in word:
            temp.resolution = 480
            return temp
        elif "720" in word:
            temp.resolution = 720
            return temp
        elif "1080" in word:
            temp.resolution = 1080
            return temp
    temp.resolition=480
    return temp

# clean up function
def clean_name(my_file):
    new_name=""
    if " " in my_file:
        split=my_file.replace('.',' ').split(' ')
    else:
        split=my_file.split('.')
    for word in split:
        if "[" in word.lower() or "]" in word.lower() or "proper" in word.lower():
            continue
        if "hdtv" in word.lower() or "web" in word.lower() or "webrip" in word.lower():
            new_name=new_name+'480P.'
            return new_name.title()
        new_name=new_name+word+'.'
        if "720" in word or "1080" in word or "480" in word:
            return new_name.title()

# create a class for our shows
class Show:
    def __init__(self,show_name,show_id,last_season,last_episode,min_res,max_res,entry,comment,time):
        self.name = show_name
        self.id = show_id
        self.season = last_season
        self.episode = last_episode
        self.min_resolution = min_res
        self.max_resolution = max_res
        self.xls_entry = entry
        self.comments=comment
        self.timestamp=time

# check xls to see what shows to search for and build them into class Show
# open the excel file, find my worksheet, and find the last used row in the sheet
wb=openpyxl.load_workbook(show_xls)
wb.sheetnames
sheet=wb[wksheet_name] #set worksheet name
last_row=sheet.max_row+1 #set last row

shows=[] #create blank list to append Show objects in
for i in range(2,last_row): #read the file and pull each column value to populate the list with objects
    # show = 'show'+str(i)
    name = sheet.cell(row=i,column=1).value
    id = sheet.cell(row=i,column=2).value
    season = sheet.cell(row=i,column=3).value
    episode = sheet.cell(row=i,column=4).value
    minres = sheet.cell(row=i,column=5).value
    maxres = sheet.cell(row=i,column=6).value
    comments = sheet.cell(row=i,column=7).value
    timestamp = sheet.cell(row=i,column=8).value
    shows.append(Show(name,id,season,episode,minres,maxres,i,comments,timestamp))

# once we have all shows built, scan to see what is currently in show show_directory to see if anything is newer
# scan directory to see what we have
file_names=[] # use this to scan against
with os.scandir(show_directory) as it:
    for each in it:
        if each.name.endswith(".mkv") and each.is_file()==True:
            file_names.append(each.name)
        elif each.name.endswith(".mp4") and each.is_file()==True:
            file_names.append(each.name)
os.scandir(show_directory).close()

# compare what is in directory and update the xls if anything newer is found
for each in shows:
    name=each.name.replace(' ','.')
    for show in file_names:
        if name.lower() in show.lower():
            info=get_ep_info(show)
            try:
                if each.season == None or info.season>each.season:
                    print(f'Updating {name} to Season: {info.season} Episode: {info.episode}')
                    sheet.cell(row=each.xls_entry,column=3).value=info.season
                    sheet.cell(row=each.xls_entry,column=4).value=info.episode
                elif info.episode>each.episode and info.season==each.season:
                    print(f'Updating {name} to Episode {info.episode}')
                    sheet.cell(row=each.xls_entry,column=4).value=info.episode
                if each.min_resolution==None:
                    print(f'Updating {name} to {info.resolution}P resolution')
                    sheet.cell(row=each.xls_entry,column=5).value=info.resolution
                    sheet.cell(row=each.xls_entry,column=6).value=info.resolution
            except ValueError:
                continue
#save the workbook
wb.save(show_xls)

# we need to update or validate any blank data in the excel sheet. these shows will not attempt to download on this phase if True.
for each in shows:
    if each.episode==None or each.season==None:
        each.update=True
    else:
        each.update=False

# once we check directory, check website for newer and compare to desired video resolution
# once all checks out, download and open torrent.
# once download starts, update csv

reserved=[] # need to keep a reserve list of files we are currently trying to download to make sure we dont delete them until script is ran again
updates=False
for show in shows:
    episode_count=0
    if show.id!=None:
        link_name = show.name.replace(" ",".").title()
        page_name=  show.name.replace(" ","-").lower()
        full_page=web_page+str(show.id)+"/"+page_name+"/"
        page=req.get(full_page)
        if page.status_code == 200:
            # get .torrent
            if show.update==True:
                soup=BeautifulSoup(page.text,'html.parser')
                link = soup.find('a', {u'class': 'download_1'})
                download = link.get('href')
                new_show = get_ep_info(str(download).split("torrent/",1)[1])
                #let's just update it in the excel sheet
                print(f'Filling out new info for {show.name} to Season: {new_show.season} Episode: {new_show.episode}')
                sheet.cell(row=show.xls_entry,column=3).value=new_show.season
                sheet.cell(row=show.xls_entry,column=4).value=new_show.episode
                wb.save(show_xls)
                continue
            else:
                soup=BeautifulSoup(page.text,'html.parser')
                downloads = soup.find_all('a',{u'class': 'download_1'})
                episodes=[]
                for link in downloads:
                    download=link.get('href')
                    if "torrents" in str(download).lower():
                        new_show = get_ep_info(str(download).split("torrents/",1)[1])
                    else:
                        new_show = get_ep_info(str(download).split("torrent/",1)[1]) 
                    epi=str(new_show.season)+str(new_show.episode)
                    if new_show.season>show.season or (new_show.season==show.season and \
                       new_show.episode>show.episode) and new_show.resolution>=show.min_resolution and \
                       new_show.resolution<=show.max_resolution and epi not in episodes and episode_count==0:
                        episodes.append(epi)
                        print(f'Downloading and updating info for {show.name} - Season: {new_show.season} Episode: {new_show.episode}')
                        sheet.cell(row=show.xls_entry,column=3).value=new_show.season
                        sheet.cell(row=show.xls_entry,column=4).value=new_show.episode
                        wb.save(show_xls)
                        if show.comments is not None:
                            if only_get_last in show.comments:
                                episode_count=1
                        # then download the .torrent
                        urlopen = urllib.request.URLopener()
                        urlopen.addheaders=[('User-Agent' , 'Mozilla/5.0')]
                        download_name = show_directory + str(download).split("torrent/",1)[1]
                        urlopen.retrieve(download , download_name)
                        # run the .torrent
                        reserved.append(download_name) # reserve the file so we dont delete it in the clean up
                        process = psutil.pids()
                        found=False
                        for id in process:
                            proc = psutil.Process(id)
                            if proc.name() == "BitTorrent.exe": # add here your process name
                                found=True
                        if found==False:
                            os.startfile(download_name) # this needs to be enabled to open file if you dont have that option turned on within BitTorrent client.
                        #email the episode info
                        subj = f'{show.name} Season: {new_show.season} Episode: {new_show.episode} now available'
                        mess = f'{show.name} Season: {new_show.season} Episode: {new_show.episode} now available'
                        email_this(subj,mess)
                        show.season=new_show.season
                        show.episode=new_show.episode
                        updates=True
if updates==False:
    print("No new episodes found at this time.")
# clean up
with os.scandir(show_directory) as it:
    for each in it:
        if each.name.endswith(".mkv") and each.is_file()==True:
            new=clean_name(each.name)
            new = str(new) + "mkv"
            os.rename(show_directory+each.name,show_directory+new)
        elif each.name.endswith(".mp4") and each.is_file()==True:
            new=clean_name(each.name)
            new = new + "mp4"
            os.rename(show_directory+each.name,show_directory+new)
        elif each.name.endswith(".torrent") and each.is_file()==True:
            if each.name not in reserved:
                print(f'Cleaning up, removing: {each.name}')
                os.remove(show_directory+ each.name)
os.scandir(show_directory).close()



    ## get magnet (could be useful later)
    ##link = soup.find('magnet')
    ##episode_name = soup.find('a', {u'class': 'epinfo'})
    ##epinfo = episode_name.get('class')
    ##magnet = link.get('href')
